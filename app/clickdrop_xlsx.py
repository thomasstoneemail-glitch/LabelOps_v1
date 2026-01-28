"""Click & Drop XLSX generation using a headerless template."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger(__name__)

ADDRESS_FIELDS = {
    "full_name",
    "address_line_1",
    "address_line_2",
    "town_city",
    "county",
    "postcode",
    "country",
}

DEFAULT_MAPPING: Dict[str, int] = {
    "full_name": 1,
    "address_line_1": 2,
    "address_line_2": 3,
    "town_city": 4,
    "county": 5,
    "postcode": 6,
    "country": 7,
    "service": 8,
    "weight_kg": 9,
    "reference": 10,
}

DEFAULTS: Dict[str, Any] = {
    "service": "T24",
    "weight_kg": 1.0,
    "country": "UNITED KINGDOM",
    "reference_prefix": None,
}


def _normalize_text(value: Any, *, uppercase: bool = False) -> str:
    """Normalize text values with trimming and optional uppercase."""
    if value is None:
        return ""
    text = str(value).strip()
    return text.upper() if uppercase else text


def _ensure_xlsx_extension(path: Path) -> Path:
    if path.suffix.lower() != ".xlsx":
        return path.with_suffix(".xlsx")
    return path


def _validate_mapping(sheet: Worksheet, mapping: Dict[str, int]) -> None:
    max_column = max(mapping.values())
    if sheet.max_column < max_column:
        LOGGER.warning(
            "Template has %s columns but mapping expects %s. Check the template.",
            sheet.max_column,
            max_column,
        )

    sample_rows = min(sheet.max_row, 5)
    if sample_rows == 0:
        LOGGER.warning("Template worksheet appears to be empty.")
        return

    populated_cells = 0
    total_cells = sample_rows * len(mapping)
    for row in range(1, sample_rows + 1):
        for column in mapping.values():
            value = sheet.cell(row=row, column=column).value
            if value not in (None, ""):
                populated_cells += 1

    if total_cells > 0 and populated_cells / total_cells < 0.1:
        LOGGER.warning(
            "Template appears mostly blank in the first %s rows. Verify mapping.",
            sample_rows,
        )


def _find_first_empty_row(sheet: Worksheet, columns: Iterable[int]) -> int:
    columns = list(columns)
    max_row = max(sheet.max_row, 1)
    for row in range(1, max_row + 2):
        if all(sheet.cell(row=row, column=column).value in (None, "") for column in columns):
            return row
    return max_row + 1


def _merge_defaults(records: list[dict], defaults: Dict[str, Any]) -> list[dict]:
    merged: list[dict] = []
    for record in records:
        merged_record = dict(record)
        for key, value in defaults.items():
            if key not in merged_record or merged_record[key] in (None, ""):
                if key != "reference_prefix":
                    merged_record[key] = value
        merged.append(merged_record)
    return merged


def _apply_reference_prefix(records: list[dict], reference_prefix: str | None) -> None:
    if not reference_prefix:
        return
    for index, record in enumerate(records, start=1):
        if record.get("reference"):
            continue
        record["reference"] = f"{reference_prefix}{index}"


def generate_clickdrop_xlsx(
    records: list[dict],
    out_path: str,
    template_path: str = r"D:\LabelOps\assets\ClickDrop_import_template_no_header.xlsx",
    *,
    defaults: dict | None = None,
) -> str:
    """
    Generate a Click & Drop XLSX using a headerless template.

    Args:
        records: Parsed address records.
        out_path: Output path to save the XLSX file.
        template_path: Path to the headerless Click & Drop template.
        defaults: Optional defaults for service, weight_kg, country, reference_prefix.

    Returns:
        The saved XLSX path as a string.
    """
    if not records:
        raise ValueError("No records provided to generate_clickdrop_xlsx.")

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(
            "Click & Drop template not found. Place the headerless template at: "
            f"{template_path}"
        )

    resolved_defaults = {**DEFAULTS, **(defaults or {})}
    records_with_defaults = _merge_defaults(records, resolved_defaults)
    _apply_reference_prefix(records_with_defaults, resolved_defaults.get("reference_prefix"))

    workbook = load_workbook(template)
    sheet = workbook.worksheets[0]

    _validate_mapping(sheet, DEFAULT_MAPPING)
    first_row = _find_first_empty_row(sheet, DEFAULT_MAPPING.values())

    for offset, record in enumerate(records_with_defaults):
        row = first_row + offset
        for field, column in DEFAULT_MAPPING.items():
            value = record.get(field)
            if field in ADDRESS_FIELDS:
                uppercase = field in {"postcode", "country"}
                value = _normalize_text(value, uppercase=uppercase)
            elif field == "service":
                value = _normalize_text(value)
            elif field == "weight_kg":
                value = float(value) if value not in (None, "") else resolved_defaults["weight_kg"]
            elif field == "reference":
                value = _normalize_text(value)
            sheet.cell(row=row, column=column, value=value)

    output_path = _ensure_xlsx_extension(Path(out_path))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    LOGGER.info("Wrote Click & Drop XLSX to %s", output_path)
    return str(output_path)


if __name__ == "__main__":
    import json

    from app.address_parser import parse_batch

    logging.basicConfig(level=logging.INFO)

    sample_input = """
Grace O'Neil
Flat 2, 10 High Street
Stonehaven
Aberdeenshire
AB538HY
UK

Martin Wilkie
Unit 7, Riverside Estate,
Dock Road
Barry
CF644BU
United Kingdom

Jamie
1 Queen's Road, Suite 5
ME74NN

James Hannay
PO Box 12
Sa198pq
Wales

M taylor
10 The Grove
Bromley
BR5 4AR

IAIN FRENCH
2 Church Lane
St Clears
Carmarthenshire
SA198PQ
""".strip()

    parsed_records = parse_batch(sample_input)
    output_file = generate_clickdrop_xlsx(
        parsed_records,
        r"D:\LabelOps\_demo_out\clickdrop_import_no_header.xlsx",
    )

    print(json.dumps(parsed_records, indent=2))
    print(f"Wrote {output_file} with {len(parsed_records)} rows")
